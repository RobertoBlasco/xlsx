
import xml.etree.ElementTree as ET
import os
import sys
import logging
import tempfile
import base64
import urllib.request
import re
from datetime import datetime
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def detect_xml_encoding(xml_file):
    """Detecta el encoding de un archivo XML basándose en su declaración"""
    try:
        # Leer las primeras líneas en modo binario para evitar errores de encoding
        with open(xml_file, 'rb') as f:
            # Leer suficientes bytes para capturar la declaración XML
            first_bytes = f.read(200)
            first_line = first_bytes.decode('ascii', errors='ignore')

            # Buscar la declaración de encoding
            import re
            encoding_match = re.search(r'encoding=["\']([^"\']+)["\']', first_line, re.IGNORECASE)
            if encoding_match:
                detected_encoding = encoding_match.group(1)
                print(f"Encoding detectado en declaración XML: {detected_encoding}")
                return detected_encoding
            else:
                print("No se encontró declaración de encoding, usando UTF-8 por defecto")
                return 'utf-8'
    except Exception as e:
        print(f"Error detectando encoding: {e}, usando UTF-8 por defecto")
        return 'utf-8'

def validate_xml_against_xsd(xml_file, xsd_file="schema.xsd"):
    """Valida el archivo XML contra el esquema XSD"""
    try:
        # Obtener la ruta del directorio donde está el script
        script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xsd_path = os.path.join(script_dir, xsd_file)

        if not os.path.exists(xsd_path):
            print(f"Advertencia: No se encontró el archivo XSD en {xsd_path}")
            return True  # Continuar sin validación

        # Cargar el esquema XSD
        with open(xsd_path, 'r', encoding='utf-8') as schema_file:
            schema_doc = etree.parse(schema_file)
            schema = etree.XMLSchema(schema_doc)

        # Detectar encoding del archivo XML y cargarlo
        xml_encoding = detect_xml_encoding(xml_file)
        with open(xml_file, 'r', encoding=xml_encoding) as xml_file_handle:
            xml_doc = etree.parse(xml_file_handle)
            
        if schema.validate(xml_doc):
            print(f"XML válido según el esquema XSD")
            return True
        else:
            print("Error: El archivo XML no es válido según el esquema XSD:")
            for error in schema.error_log:
                print(f"  Línea {error.line}: {error.message}")
            return False
            
    except etree.XMLSyntaxError as e:
        print(f"Error de sintaxis XML: {e}")
        return False
    except Exception as e:
        print(f"Error durante la validación XSD: {e}")
        return False

def setup_logging(log_element):
    """Configura el sistema de logging según la configuración XML"""
    if log_element is None:
        # Configuración por defecto
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        return logging.getLogger(__name__)
    
    # Obtener configuraciones
    log_level_elem = log_element.find('logLevel')
    log_file_elem = log_element.find('logFile')
    log_format_elem = log_element.find('logFormat')
    log_date_format_elem = log_element.find('logDateFormat')
    log_console_elem = log_element.find('logConsole')
    
    # Configurar nivel de log con validación
    log_level_text = log_level_elem.text.upper() if log_level_elem is not None else 'INFO'
    try:
        level = getattr(logging, log_level_text)
    except AttributeError:
        level = logging.INFO
        print(f"Advertencia: Nivel de log '{log_level_text}' no válido, usando INFO")
    
    # Configurar formato
    log_format = log_format_elem.text if log_format_elem is not None else '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    date_format = log_date_format_elem.text if log_date_format_elem is not None else '%Y-%m-%d %H:%M:%S'
    
    # Configurar handlers
    handlers = []
    
    # Handler para archivo si está especificado
    if log_file_elem is not None:
        try:
            # Extraer ruta del archivo (manejar prefijo FILE://)
            log_file_text = log_file_elem.text
            if log_file_text.startswith('FILE://'):
                log_file_path = log_file_text[7:]
            else:
                log_file_path = log_file_text
            
            # Crear directorio si no existe
            log_dir = os.path.dirname(log_file_path)
            if log_dir and not os.path.exists(log_dir):
                os.makedirs(log_dir, exist_ok=True)
            
            file_handler = logging.FileHandler(log_file_path, encoding='utf-8')
            file_handler.setFormatter(logging.Formatter(log_format, date_format))
            handlers.append(file_handler)
        except Exception as e:
            print(f"Advertencia: No se pudo configurar el archivo de log: {e}")
    
    # Handler para consola si está habilitado
    log_console_text = log_console_elem.text.lower() if log_console_elem is not None else 'true'
    log_console = log_console_text in ['true', '1', 'yes', 'on']
    
    if log_console:
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(logging.Formatter(log_format, date_format))
        handlers.append(console_handler)
    
    # Si no hay handlers, agregar uno por defecto
    if not handlers:
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(logging.Formatter(log_format, date_format))
        handlers.append(console_handler)
    
    # Configurar logger con nombre único para evitar conflictos
    logger_name = f"{__name__}_{id(log_element) if log_element is not None else 'default'}"
    logger = logging.getLogger(logger_name)
    logger.setLevel(level)
    
    # Limpiar handlers existentes
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    # Agregar nuevos handlers
    for handler in handlers:
        logger.addHandler(handler)
    
    # Evitar propagación para evitar logs duplicados
    logger.propagate = False
    
    return logger

def extract_uri_content(uri_string):
    """Extrae el tipo y contenido de una URI con prefijos FILE://, BASE64://, URL://"""
    if not uri_string:
        return 'file', uri_string
    
    if uri_string.startswith('FILE://'):
        return 'file', uri_string[7:]
    elif uri_string.startswith('BASE64://'):
        return 'base64', uri_string[9:]
    elif uri_string.startswith('URL://'):
        return 'url', uri_string[6:]
    else:
        # Por defecto es archivo
        return 'file', uri_string

def validate_and_get_data_source(uri_string, logger=None, is_data_in=True):
    """Valida y obtiene el archivo de datos según el prefijo URI"""
    try:
        uri_type, content = extract_uri_content(uri_string)
        
        if uri_type == 'file':
            # Para dataIn, validar que el archivo existe
            if is_data_in and not os.path.exists(content):
                if logger:
                    logger.error(f"El archivo de entrada no existe: {content}")
                else:
                    print(f"Error: El archivo de entrada no existe: {content}")
                return None
            
            # Para dataOut, crear directorio si no existe
            if not is_data_in:
                output_dir = os.path.dirname(content)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir, exist_ok=True)
                    if logger:
                        logger.info(f"Directorio creado: {output_dir}")
                    else:
                        print(f"Directorio creado: {output_dir}")
            
            return content
            
        elif uri_type == 'base64':
            if not is_data_in:
                if logger:
                    logger.error("BASE64:// no es válido para dataOut")
                else:
                    print("Error: BASE64:// no es válido para dataOut")
                return None
                
            try:
                # Decodificar BASE64 y crear archivo temporal
                decoded_content = base64.b64decode(content).decode('utf-8')
                temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xml', delete=False, encoding='utf-8')
                temp_file.write(decoded_content)
                temp_file.close()
                
                if logger:
                    logger.info(f"Contenido BASE64 decodificado a archivo temporal: {temp_file.name}")
                
                return temp_file.name
                
            except Exception as e:
                if logger:
                    logger.error(f"Error decodificando BASE64: {e}")
                else:
                    print(f"Error decodificando BASE64: {e}")
                return None
                
        elif uri_type == 'url':
            if is_data_in:
                # Descargar desde URL para dataIn
                try:
                    response = urllib.request.urlopen(content)
                    temp_file = tempfile.NamedTemporaryFile(mode='wb', suffix='.xml', delete=False)
                    temp_file.write(response.read())
                    temp_file.close()
                    
                    if logger:
                        logger.info(f"Archivo descargado desde URL: {content} -> {temp_file.name}")
                    
                    return temp_file.name
                    
                except Exception as e:
                    if logger:
                        logger.error(f"Error descargando desde URL: {e}")
                    else:
                        print(f"Error descargando desde URL: {e}")
                    return None
            else:
                # Para dataOut, URL:// no está implementado aún
                if logger:
                    logger.error("URL:// para dataOut no está implementado")
                else:
                    print("Error: URL:// para dataOut no está implementado")
                return None
                
    except Exception as e:
        if logger:
            logger.error(f"Error validando origen de datos: {e}")
        else:
            print(f"Error validando origen de datos: {e}")
        return None

def parse_styles(styles_element):
    """Parsea los estilos del XML y retorna un diccionario"""
    styles_dict = {}
    for style in styles_element.findall('style'):
        style_id = style.get('id')
        style_data = {}
        font_elem = style.find('font')
        if font_elem is not None:
            style_data['font'] = font_elem.text
        size_elem = style.find('size')
        if size_elem is not None:
            style_data['size'] = int(size_elem.text)
        bold_elem = style.find('bold')
        if bold_elem is not None:
            style_data['bold'] = bold_elem.text.lower() == 'true'
        italic_elem = style.find('italic')
        if italic_elem is not None:
            style_data['italic'] = italic_elem.text.lower() == 'true'
        underline_elem = style.find('underline')
        if underline_elem is not None:
            style_data['underline'] = underline_elem.text.lower() == 'true'
        strikethrough_elem = style.find('strikethrough')
        if strikethrough_elem is not None:
            style_data['strikethrough'] = strikethrough_elem.text.lower() == 'true'
        color_elem = style.find('color')
        if color_elem is not None:
            style_data['color'] = color_elem.text.replace('#', '')
        background_elem = style.find('background')
        if background_elem is not None:
            style_data['background'] = background_elem.text.replace('#', '')
        # Alignment removido - ahora se maneja como atributo de celda
        styles_dict[style_id] = style_data
    return styles_dict

def create_openpyxl_style(style_data):
    """Convierte los datos de estilo a objetos openpyxl"""
    font = Font(
        name=style_data.get('font', 'Arial'),
        size=style_data.get('size', 10),
        bold=style_data.get('bold', False),
        italic=style_data.get('italic', False),
        underline='single' if style_data.get('underline', False) else 'none',
        strike=style_data.get('strikethrough', False),
        color=style_data.get('color', '000000')
    )
    fill = None
    if 'background' in style_data:
        fill = PatternFill(
            start_color=style_data['background'],
            end_color=style_data['background'],
            fill_type='solid'
        )
    # Alignment removido de estilos - ahora se maneja como atributo de celda
    return font, fill, None

def parse_column_settings(workbook_elem):
    """Parsea las configuraciones de columna del XML"""
    column_settings = {}
    column_settings_elem = workbook_elem.find('columnSettings')
    if column_settings_elem is not None:
        for column_elem in column_settings_elem.findall('column'):
            column_name = column_elem.get('name')
            column_config = {
                'width': column_elem.get('width'),
                'format': column_elem.get('format'),
                'style': column_elem.get('style'),
                'horizontalAlignment': column_elem.get('horizontalAlignment'),
                'verticalAlignment': column_elem.get('verticalAlignment')
            }
            # Filtrar valores None
            column_config = {k: v for k, v in column_config.items() if v is not None}
            column_settings[column_name] = column_config
    return column_settings

def parse_row_settings(workbook_elem):
    """Parsea las configuraciones de fila del XML"""
    row_settings = {}
    row_settings_elem = workbook_elem.find('rowSettings')
    if row_settings_elem is not None:
        for row_elem in row_settings_elem.findall('row'):
            row_number = int(row_elem.get('number'))
            row_config = {
                'height': row_elem.get('height'),
                'format': row_elem.get('format'),
                'style': row_elem.get('style'),
                'horizontalAlignment': row_elem.get('horizontalAlignment'),
                'verticalAlignment': row_elem.get('verticalAlignment')
            }
            # Filtrar valores None
            row_config = {k: v for k, v in row_config.items() if v is not None}
            row_settings[row_number] = row_config
    return row_settings

def get_cell_properties(row, column, cell_elem, column_settings, row_settings):
    """Obtiene las propiedades finales de una celda aplicando precedencia"""
    # Inicializar con valores por defecto
    properties = {
        'width': None,
        'height': None,
        'format': 'General',
        'style': None,
        'horizontalAlignment': None,
        'verticalAlignment': None
    }
    
    # 1. Aplicar configuración de columna (prioridad baja)
    if column in column_settings:
        properties.update({k: v for k, v in column_settings[column].items() if v is not None})
    
    # 2. Aplicar configuración de fila (prioridad media)
    if row in row_settings:
        properties.update({k: v for k, v in row_settings[row].items() if v is not None})
    
    # 3. Aplicar propiedades específicas de celda (prioridad máxima)
    cell_properties = {
        'width': cell_elem.get('width'),
        'format': cell_elem.get('format'),
        'style': cell_elem.get('style'),
        'horizontalAlignment': cell_elem.get('horizontalAlignment'),
        'verticalAlignment': cell_elem.get('verticalAlignment')
    }
    properties.update({k: v for k, v in cell_properties.items() if v is not None})
    
    return properties

def parse_tables(workbook_elem):
    """Parsea las tablas del XML"""
    tables = []
    table_elements = workbook_elem.findall('table')
    for table_elem in table_elements:
        table_config = {
            'name': table_elem.get('name'),
            'ref': table_elem.get('ref'),
            'style': table_elem.get('style', 'TableStyleMedium9'),
            'showFirstColumn': table_elem.get('showFirstColumn', 'false').lower() == 'true',
            'showLastColumn': table_elem.get('showLastColumn', 'false').lower() == 'true',
            'showRowStripes': table_elem.get('showRowStripes', 'true').lower() == 'true',
            'showColumnStripes': table_elem.get('showColumnStripes', 'false').lower() == 'true'
        }
        tables.append(table_config)
    return tables

def parse_options(root_elem):
    """Parsea las opciones globales del XML"""
    options = {
        'autoAdjustColumnWidth': True,  # Por defecto activado (comportamiento actual)
        'maxColumnWidth': 50,
        'minColumnWidth': 8
    }

    options_elem = root_elem.find('options')
    if options_elem is not None:
        for option_elem in options_elem.findall('option'):
            name = option_elem.get('name')
            value = option_elem.get('value')

            if name == 'autoAdjustColumnWidth':
                options['autoAdjustColumnWidth'] = value.lower() == 'true'
            elif name == 'maxColumnWidth':
                try:
                    options['maxColumnWidth'] = int(value)
                except ValueError:
                    pass  # Mantener valor por defecto si no es válido
            elif name == 'minColumnWidth':
                try:
                    options['minColumnWidth'] = int(value)
                except ValueError:
                    pass  # Mantener valor por defecto si no es válido

    return options

def convert_value_by_format(value_str, number_format, logger=None):
    """
    Convierte el valor string al tipo apropiado según el formato especificado.

    Parámetros:
        value_str: Valor como string del XML
        number_format: Formato de número/fecha (ej: 'dd/mm/yyyy', '0', '€#,##0.00')
        logger: Logger opcional para debug

    Retorna:
        El valor convertido al tipo apropiado (datetime, int, float, o str)
    """
    if not value_str or value_str.strip() == '':
        return value_str

    value_str = value_str.strip()

    # Detectar si es un formato de fecha
    date_format_patterns = [
        r'd+[/-]m+[/-]y+',  # dd/mm/yyyy, dd-mm-yyyy, d/m/yy, etc.
        r'm+[/-]d+[/-]y+',  # mm/dd/yyyy, mm-dd-yyyy, m/d/yy, etc.
        r'y+[/-]m+[/-]d+',  # yyyy/mm/dd, yyyy-mm-dd, yy/m/d, etc.
        r'd+\.m+\.y+',      # dd.mm.yyyy, d.m.yy, etc.
    ]

    is_date_format = any(re.search(pattern, number_format.lower()) for pattern in date_format_patterns)

    if is_date_format:
        # Intentar parsear como fecha
        # Formatos de fecha comunes que puede recibir como valor
        date_input_formats = [
            '%Y-%m-%d',           # 2023-01-15
            '%d/%m/%Y',           # 15/01/2023
            '%d-%m-%Y',           # 15-01-2023
            '%d.%m.%Y',           # 15.01.2023
            '%m/%d/%Y',           # 01/15/2023
            '%Y/%m/%d',           # 2023/01/15
            '%Y%m%d',             # 20230115
            '%d/%m/%y',           # 15/01/23
            '%d-%m-%y',           # 15-01-23
        ]

        for date_format in date_input_formats:
            try:
                parsed_date = datetime.strptime(value_str, date_format)
                if logger:
                    logger.debug(f"Valor '{value_str}' convertido a fecha usando formato '{date_format}'")
                return parsed_date
            except ValueError:
                continue

        # Si no se pudo parsear como fecha, advertir y retornar como string
        if logger:
            logger.warning(f"No se pudo convertir '{value_str}' a fecha. Formatos intentados: {date_input_formats}. Se guardará como texto.")
        return value_str

    # Detectar si es un formato numérico (pero no fecha)
    # Patrones de formato numérico: 0, 0.00, #,##0, #,##0.00, €#,##0.00, 0.00%, etc.
    numeric_format_patterns = [
        r'^[#0]',              # Empieza con # o 0
        r'[€$£¥]',             # Símbolos de moneda
        r'%',                  # Porcentaje
        r'#,##0',              # Separador de miles
    ]

    is_numeric_format = any(re.search(pattern, number_format) for pattern in numeric_format_patterns)

    if is_numeric_format:
        # Intentar convertir a número
        try:
            # Limpiar el valor de posibles separadores de miles y espacios
            clean_value = value_str.replace(',', '').replace(' ', '').replace('\xa0', '')

            # Intentar como float
            if '.' in clean_value or 'e' in clean_value.lower():
                numeric_value = float(clean_value)
                if logger:
                    logger.debug(f"Valor '{value_str}' convertido a float: {numeric_value}")
                return numeric_value
            else:
                # Intentar como int
                numeric_value = int(clean_value)
                if logger:
                    logger.debug(f"Valor '{value_str}' convertido a int: {numeric_value}")
                return numeric_value
        except ValueError:
            # Si no se puede convertir, advertir y retornar como string
            if logger:
                logger.warning(f"No se pudo convertir '{value_str}' a número según formato '{number_format}'. Se guardará como texto.")
            return value_str

    # Si no es fecha ni número, retornar como string
    return value_str

def xml_to_excel(config_file, output_file=None):
    """Convierte el XML a Excel usando configuración del archivo"""
    logger = None
    
    try:
        # Validar XML contra XSD antes de procesarlo
        if not validate_xml_against_xsd(config_file):
            return False
        
        tree = ET.parse(config_file)
        root = tree.getroot()
        
        # Configurar logging primero
        log_element = root.find('log')
        logger = setup_logging(log_element)
        logger.info("Iniciando conversión XML a Excel")
        
        # Log de la configuración de logging
        if log_element is not None:
            log_config = {}
            for child in log_element:
                log_config[child.tag] = child.text
            logger.info(f"Configuración de logging aplicada: {log_config}")
        else:
            logger.info("No se especificó configuración de logging, usando configuración por defecto")
        
        # Extraer configuración de datos
        data_element = root.find('data')
        if data_element is not None:
            data_in_element = data_element.find('dataIn')
            data_out_element = data_element.find('dataOut')
            
            if data_in_element is not None:
                logger.info(f"dataIn especificado: {data_in_element.text}")
                xml_file = validate_and_get_data_source(data_in_element.text, logger, is_data_in=True)
                if xml_file is None:
                    return False
                logger.info(f"Archivo XML de datos procesado: {xml_file}")
            else:
                # Si no hay dataIn, usar el archivo de configuración como datos
                xml_file = config_file
                logger.info("No se especificó dataIn, usando archivo de configuración como datos")
                
            if data_out_element is not None:
                excel_file = validate_and_get_data_source(data_out_element.text, logger, is_data_in=False)
                if excel_file is None:
                    return False
                logger.info(f"Archivo Excel de salida: {excel_file}")
            elif output_file:
                excel_file = output_file
                logger.info(f"Usando archivo de salida del parámetro: {excel_file}")
            else:
                logger.error("No se encontró dataOut en la configuración ni se especificó archivo de salida")
                return False
        else:
            # Fallback: usar el archivo de configuración como XML de datos
            xml_file = config_file
            excel_file = output_file if output_file else "salida.xlsx"
            if logger:
                logger.warning("No se encontró sección <data>, usando modo compatibilidad")
            else:
                print("Advertencia: No se encontró sección <data>, usando modo compatibilidad")
        
        # Si xml_file es diferente al config_file, cargar el archivo de datos
        if xml_file != config_file:
            if not os.path.exists(xml_file):
                print(f"Error: El archivo de datos {xml_file} no existe")
                return False
            data_tree = ET.parse(xml_file)
            data_root = data_tree.getroot()
        else:
            data_root = root
        
        # Buscar estilos y workbooks en el archivo de datos
        styles_element = data_root.find('styles')
        if styles_element is None:
            # Si no hay estilos en datos, buscar en workbooks
            workbooks_element = data_root.find('workbooks')
            if workbooks_element is not None:
                styles_element = workbooks_element.find('styles')
        
        styles_dict = parse_styles(styles_element) if styles_element is not None else {}
        
        # Parsear opciones globales
        global_options = parse_options(root)
        
        # Verificar si el archivo Excel ya existe
        if os.path.isfile(excel_file):
            wb = load_workbook(excel_file)
            print(f"Cargando archivo Excel existente: {excel_file}")
        else:
            wb = Workbook()
            wb.remove(wb.active)
            print(f"Creando nuevo archivo Excel: {excel_file}")
        
        created_sheets = {}  # Diccionario para rastrear hojas creadas
        
        # Si se cargó un archivo existente, indexar las hojas ya presentes
        for existing_sheet in wb.worksheets:
            created_sheets[existing_sheet.title] = existing_sheet
        
        # Buscar workbooks en el lugar correcto
        workbook_elements = data_root.findall('workbook')
        if not workbook_elements:
            # Si no hay workbooks directos, buscar en sección workbooks
            workbooks_element = data_root.find('workbooks')
            if workbooks_element is not None:
                workbook_elements = workbooks_element.findall('workbook')
        
        for workbook_elem in workbook_elements:
            sheet_name = workbook_elem.get('name', 'Hoja1')
            
            # Determinar configuración de autoAdjustColumnWidth para esta hoja
            # Precedencia: workbook específico > global
            workbook_auto_adjust = workbook_elem.get('autoAdjustColumnWidth')
            if workbook_auto_adjust is not None:
                auto_adjust_enabled = workbook_auto_adjust.lower() == 'true'
            else:
                auto_adjust_enabled = global_options['autoAdjustColumnWidth']
            
            # Parsear configuraciones de columna y fila
            column_settings = parse_column_settings(workbook_elem)
            row_settings = parse_row_settings(workbook_elem)
            tables = parse_tables(workbook_elem)
            
            # Contar las celdas en este workbook
            cell_elements = workbook_elem.findall('cell')
            cell_count = len(cell_elements)
            
            # Verificar si la hoja ya existe (en archivo cargado o ya procesada)
            if sheet_name in created_sheets:
                ws = created_sheets[sheet_name]
                print(f"  Utilizando hoja existente: {sheet_name}")
                if logger:
                    logger.info(f"Workbook '{sheet_name}': Utilizando hoja existente, procesando {cell_count} celdas")
            else:
                ws = wb.create_sheet(title=sheet_name)
                created_sheets[sheet_name] = ws
                print(f"  Creando nueva hoja: {sheet_name}")
                if logger:
                    logger.info(f"Workbook '{sheet_name}': Hoja creada, procesando {cell_count} celdas")
            
            # Log configuraciones si existen
            if column_settings and logger:
                logger.info(f"Configuraciones de columna aplicadas: {list(column_settings.keys())}")
            if row_settings and logger:
                logger.info(f"Configuraciones de fila aplicadas: {list(row_settings.keys())}")
            
            for cell_elem in cell_elements:
                row = int(cell_elem.get('row'))
                column = cell_elem.get('column')
                value = cell_elem.get('value', '')

                # Obtener propiedades finales aplicando precedencia
                properties = get_cell_properties(row, column, cell_elem, column_settings, row_settings)

                # Convertir el valor según el formato antes de asignarlo
                converted_value = convert_value_by_format(value, properties['format'], logger)

                ws[f"{column}{row}"] = converted_value
                cell = ws[f"{column}{row}"]

                # Aplicar formato de número si está especificado
                if properties['format'] and properties['format'] != 'General':
                    cell.number_format = properties['format']
                    if logger:
                        logger.debug(f"Aplicando formato '{properties['format']}' a celda {column}{row}")
                
                # Aplicar width si está especificado
                if properties['width']:
                    width_value = int(properties['width'])
                    ws.column_dimensions[column].width = width_value
                    if logger:
                        logger.debug(f"Aplicando width '{width_value}' a columna {column}")
                
                # Aplicar height si está especificado
                if properties['height']:
                    height_value = int(properties['height'])
                    ws.row_dimensions[row].height = height_value
                    if logger:
                        logger.debug(f"Aplicando height '{height_value}' a fila {row}")
                
                # Aplicar alineamiento si está especificado
                if properties['horizontalAlignment'] or properties['verticalAlignment']:
                    alignment = Alignment(
                        horizontal=properties['horizontalAlignment'],
                        vertical=properties['verticalAlignment']
                    )
                    cell.alignment = alignment
                    if logger:
                        alignment_info = []
                        if properties['horizontalAlignment']:
                            alignment_info.append(f"horizontal={properties['horizontalAlignment']}")
                        if properties['verticalAlignment']:
                            alignment_info.append(f"vertical={properties['verticalAlignment']}")
                        logger.debug(f"Aplicando alineamiento {', '.join(alignment_info)} a celda {column}{row}")
                
                # Aplicar estilos si están especificados
                if properties['style'] and properties['style'] in styles_dict:
                    style_data = styles_dict[properties['style']]
                    font, fill, _ = create_openpyxl_style(style_data)
                    cell.font = font
                    if fill:
                        cell.fill = fill
            
            # Procesar tablas después de las celdas
            for table_config in tables:
                try:
                    # Crear el objeto TableStyleInfo
                    style_info = TableStyleInfo(
                        name=table_config['style'],
                        showFirstColumn=table_config['showFirstColumn'],
                        showLastColumn=table_config['showLastColumn'],
                        showRowStripes=table_config['showRowStripes'],
                        showColumnStripes=table_config['showColumnStripes']
                    )
                    
                    # Crear el objeto Table
                    table = Table(
                        displayName=table_config['name'],
                        ref=table_config['ref']
                    )
                    table.tableStyleInfo = style_info
                    
                    # Añadir la tabla a la hoja
                    ws.add_table(table)
                    
                    if logger:
                        logger.info(f"Tabla '{table_config['name']}' creada en rango {table_config['ref']} con estilo {table_config['style']}")
                        
                except Exception as e:
                    if logger:
                        logger.error(f"Error creando tabla '{table_config['name']}': {e}")
            
            # Aplicar ajuste automático de anchos si está habilitado para esta hoja
            if auto_adjust_enabled:
                columns_with_explicit_width = set()
                # Identificar columnas que tienen width explícito (no se deben ajustar)
                for column_name, config in column_settings.items():
                    if 'width' in config:
                        columns_with_explicit_width.add(column_name)
                
                for column in ws.columns:
                    column_letter = get_column_letter(column[0].column)
                    
                    # Solo ajustar si no tiene width explícito configurado
                    if column_letter not in columns_with_explicit_width:
                        max_length = 0
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        if max_length > 0:  # Solo ajustar si hay contenido
                            adjusted_width = max(
                                global_options['minColumnWidth'],
                                min(max_length + 2, global_options['maxColumnWidth'])
                            )
                            ws.column_dimensions[column_letter].width = adjusted_width
                            
                            if logger:
                                logger.debug(f"Ajuste automático columna {column_letter}: width={adjusted_width} (contenido max: {max_length})")
                
                if logger:
                    excluded_count = len(columns_with_explicit_width)
                    logger.info(f"Ajuste automático aplicado a hoja '{sheet_name}' (excluidas {excluded_count} columnas con width explícito)")
            else:
                if logger:
                    logger.info(f"Ajuste automático deshabilitado para hoja '{sheet_name}'")
        wb.save(excel_file)
        if logger:
            logger.info(f"Archivo Excel creado exitosamente: {excel_file}")
        else:
            print(f"Archivo Excel creado exitosamente: {excel_file}")
        return True
        
    except ET.ParseError as e:
        error_msg = f"Error parsing XML: {e}"
        if logger:
            logger.error(error_msg)
        else:
            print(error_msg)
        return False
    except Exception as e:
        error_msg = f"Error creando Excel: {e}"
        if logger:
            logger.error(error_msg)
        else:
            print(error_msg)
        return False