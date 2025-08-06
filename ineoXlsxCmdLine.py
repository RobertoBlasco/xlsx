#!/usr/bin/env python3
"""
Programa principal para convertir XML a Excel desde línea de comandos.
Lee el archivo XML indicado y genera salida.xlsx
"""


# TODO Añadir overwrite en dataOut al xsd

import ineoXlsxGlobales

from datetime import datetime
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time
import sys
import os
from lxml import etree
import base64
import urllib.request
import tempfile
import logging



def validate_xml_against_xsd(xml_file, xsd_file="schema.xsd"):
    """Valida el archivo XML contra el esquema XSD"""
    try:
        # Obtener la ruta del directorio donde está el script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        xsd_path = os.path.join(script_dir, xsd_file)
        
        if not os.path.exists(xsd_path):
            print(f"Advertencia: No se encontró el archivo XSD en {xsd_path}")
            return True  # Continuar sin validación
        
        # Cargar el esquema XSD
        with open(xsd_path, 'r', encoding='utf-8') as schema_file:
            schema_doc = etree.parse(schema_file)
            schema = etree.XMLSchema(schema_doc)
        
        # Cargar y validar el XML
        with open(xml_file, 'r', encoding='utf-8') as xml_file_handle:
            xml_doc = etree.parse(xml_file_handle)
            
        if schema.validate(xml_doc):
            print(f"XML válido según el esquema XSD")
            return True
        else:
            print("Error: El archivo XML no es válido según el esquema XSD:")
            for error in schema.error_log:
                print(f"  Línea {error.line}: {error.message}")
            return False
            
    except etree.XMLSyntaxError as e:
        print(f"Error de sintaxis XML: {e}")
        return False
    except Exception as e:
        print(f"Error durante la validación XSD: {e}")
        return False

def parse_styles(styles_element):
    """Parsea los estilos del XML y retorna un diccionario"""
    styles_dict = {}
    for style in styles_element.findall('style'):
        style_id = style.get('id')
        style_data = {}
        font_elem = style.find('font')
        if font_elem is not None:
            style_data['font'] = font_elem.text
        size_elem = style.find('size')
        if size_elem is not None:
            style_data['size'] = int(size_elem.text)
        bold_elem = style.find('bold')
        if bold_elem is not None:
            style_data['bold'] = bold_elem.text.lower() == 'true'
        color_elem = style.find('color')
        if color_elem is not None:
            style_data['color'] = color_elem.text.replace('#', '')
        background_elem = style.find('background')
        if background_elem is not None:
            style_data['background'] = background_elem.text.replace('#', '')
        alignment_elem = style.find('alignment')
        if alignment_elem is not None:
            style_data['alignment'] = alignment_elem.text
        styles_dict[style_id] = style_data
    return styles_dict

def create_openpyxl_style(style_data):
    """Convierte los datos de estilo a objetos openpyxl"""
    font = Font(
        name=style_data.get('font', 'Arial'),
        size=style_data.get('size', 10),
        bold=style_data.get('bold', False),
        color=style_data.get('color', '000000')
    )
    fill = None
    if 'background' in style_data:
        fill = PatternFill(
            start_color=style_data['background'],
            end_color=style_data['background'],
            fill_type='solid'
        )
    alignment = None
    if 'alignment' in style_data:
        horizontal = style_data['alignment']
        alignment = Alignment(horizontal=horizontal)
    return font, fill, alignment

def setup_logging(log_element):
    """Configura el sistema de logging según la configuración XML"""
    if log_element is None:
        # Configuración por defecto
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        return logging.getLogger(__name__)
    
    # Obtener configuraciones
    log_level_elem = log_element.find('logLevel')
    log_file_elem = log_element.find('logFile')
    log_format_elem = log_element.find('logFormat')
    log_date_format_elem = log_element.find('logDateFormat')
    log_console_elem = log_element.find('logConsole')
    
    # Configurar nivel de log con validación
    log_level_text = log_level_elem.text.upper() if log_level_elem is not None else 'INFO'
    try:
        level = getattr(logging, log_level_text)
    except AttributeError:
        level = logging.INFO
        print(f"Advertencia: Nivel de log '{log_level_text}' no válido, usando INFO")
    
    # Configurar formato
    log_format = log_format_elem.text if log_format_elem is not None else '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    date_format = log_date_format_elem.text if log_date_format_elem is not None else '%Y-%m-%d %H:%M:%S'
    
    # Configurar handlers
    handlers = []
    
    # Handler para archivo si está especificado
    if log_file_elem is not None:
        try:
            # Extraer ruta del archivo (manejar prefijo FILE://)
            log_file_text = log_file_elem.text
            if log_file_text.startswith('FILE://'):
                log_file_path = log_file_text[7:]
            else:
                log_file_path = log_file_text
            
            # Crear directorio si no existe
            log_dir = os.path.dirname(log_file_path)
            if log_dir and not os.path.exists(log_dir):
                os.makedirs(log_dir, exist_ok=True)
            
            file_handler = logging.FileHandler(log_file_path, encoding='utf-8')
            file_handler.setFormatter(logging.Formatter(log_format, date_format))
            handlers.append(file_handler)
        except Exception as e:
            print(f"Advertencia: No se pudo configurar el archivo de log: {e}")
    
    # Handler para consola si está habilitado
    log_console_text = log_console_elem.text.lower() if log_console_elem is not None else 'true'
    log_console = log_console_text in ['true', '1', 'yes', 'on']
    
    if log_console:
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(logging.Formatter(log_format, date_format))
        handlers.append(console_handler)
    
    # Si no hay handlers, agregar uno por defecto
    if not handlers:
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(logging.Formatter(log_format, date_format))
        handlers.append(console_handler)
    
    # Configurar logger con nombre único para evitar conflictos
    logger_name = f"{__name__}_{id(log_element) if log_element is not None else 'default'}"
    logger = logging.getLogger(logger_name)
    logger.setLevel(level)
    
    # Limpiar handlers existentes
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    # Agregar nuevos handlers
    for handler in handlers:
        logger.addHandler(handler)
    
    # Evitar propagación para evitar logs duplicados
    logger.propagate = False
    
    return logger

def extract_uri_content(uri_string):
    """Extrae el tipo y contenido de una URI con prefijos FILE://, BASE64://, URL://"""
    if not uri_string:
        return 'file', uri_string
    
    if uri_string.startswith('FILE://'):
        return 'file', uri_string[7:]
    elif uri_string.startswith('BASE64://'):
        return 'base64', uri_string[9:]
    elif uri_string.startswith('URL://'):
        return 'url', uri_string[6:]
    else:
        # Por defecto es archivo
        return 'file', uri_string




def main():

    ineoXlsxGlobales.EXECUTION_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")

    """Función principal"""
    if len(sys.argv) < 2:
        print("Uso: python ineoXlsxCmdLine.py <archivo_xml> [archivo_excel]")
        sys.exit(1)
    xml_file = sys.argv[1]
    if not os.path.isfile(xml_file):
        print(f"El archivo {xml_file} no existe.")
        sys.exit(1)
    excel_file = sys.argv[2] if len(sys.argv) > 2 else "salida.xlsx"
    print(f"Convirtiendo {xml_file} a {excel_file}...")
    start_time = time.time()
    print(f"Inicio: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(start_time))}")
    if xml_to_excel(xml_file, excel_file):
        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"Fin: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end_time))}")
        print(f"Tiempo transcurrido: {elapsed_time:.2f} segundos")
        print("Conversión completada exitosamente")
    else:
        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"Fin: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end_time))}")
        print(f"Tiempo transcurrido: {elapsed_time:.2f} segundos")
        print("Error en la conversión")

if __name__ == "__main__":
    main()