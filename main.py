#!/usr/bin/env python3
"""
Programa principal para convertir XML a Excel
Lee el archivo ineo_xlsx.xml y genera salida.xlsx
"""

import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time

def parse_styles(styles_element):
    """Parsea los estilos del XML y retorna un diccionario"""
    styles_dict = {}
    
    for style in styles_element.findall('style'):
        style_id = style.get('id')
        style_data = {}
        
        # Parsear propiedades del estilo
        font_elem = style.find('font')
        if font_elem is not None:
            style_data['font'] = font_elem.text
            
        size_elem = style.find('size')
        if size_elem is not None:
            style_data['size'] = int(size_elem.text)
            
        bold_elem = style.find('bold')
        if bold_elem is not None:
            style_data['bold'] = bold_elem.text.lower() == 'true'
            
        color_elem = style.find('color')
        if color_elem is not None:
            style_data['color'] = color_elem.text.replace('#', '')
            
        background_elem = style.find('background')
        if background_elem is not None:
            style_data['background'] = background_elem.text.replace('#', '')
            
        alignment_elem = style.find('alignment')
        if alignment_elem is not None:
            style_data['alignment'] = alignment_elem.text
            
        styles_dict[style_id] = style_data
    
    return styles_dict

def create_openpyxl_style(style_data):
    """Convierte los datos de estilo a objetos openpyxl"""
    font = Font(
        name=style_data.get('font', 'Arial'),
        size=style_data.get('size', 10),
        bold=style_data.get('bold', False),
        color=style_data.get('color', '000000')
    )
    
    fill = None
    if 'background' in style_data:
        fill = PatternFill(
            start_color=style_data['background'],
            end_color=style_data['background'],
            fill_type='solid'
        )
    
    alignment = None
    if 'alignment' in style_data:
        horizontal = style_data['alignment']
        alignment = Alignment(horizontal=horizontal)
    
    return font, fill, alignment

def xml_to_excel(xml_file, excel_file):
    """Convierte el XML a Excel"""
    try:
        # Parsear XML
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        # Obtener estilos
        styles_element = root.find('styles')
        styles_dict = parse_styles(styles_element) if styles_element is not None else {}
        
        # Crear workbook
        wb = Workbook()
        # Remover la hoja por defecto
        wb.remove(wb.active)
        
        # Procesar cada workbook
        for workbook_elem in root.findall('workbook'):
            sheet_name = workbook_elem.get('name', 'Hoja1')
            ws = wb.create_sheet(title=sheet_name)
            
            # Procesar cada celda
            for cell_elem in workbook_elem.findall('cell'):
                row = int(cell_elem.get('row'))
                column = cell_elem.get('column')
                text = cell_elem.get('text', '')
                style_id = cell_elem.get('style')
                
                # Escribir valor en la celda
                ws[f"{column}{row}"] = text
                
                # Aplicar estilo si existe
                if style_id and style_id in styles_dict:
                    style_data = styles_dict[style_id]
                    font, fill, alignment = create_openpyxl_style(style_data)
                    
                    cell = ws[f"{column}{row}"]
                    cell.font = font
                    if fill:
                        cell.fill = fill
                    if alignment:
                        cell.alignment = alignment
        
        # Ajustar ancho de columnas automáticamente
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo
        wb.save(excel_file)
        print(f"Archivo Excel creado exitosamente: {excel_file}")
        return True
        
    except ET.ParseError as e:
        print(f"Error parsing XML: {e}")
        return False
    except Exception as e:
        print(f"Error creando Excel: {e}")
        return False

def main():
    """Función principal"""
    xml_file = "ineo_xlsx.xml"
    excel_file = "salida.xlsx"
    
    print(f"Convirtiendo {xml_file} a {excel_file}...")
    
    # Timestamp de inicio
    start_time = time.time()
    print(f"Inicio: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(start_time))}")
    
    if xml_to_excel(xml_file, excel_file):
        # Timestamp de fin
        end_time = time.time()
        elapsed_time = end_time - start_time
        
        print(f"Fin: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end_time))}")
        print(f"Tiempo transcurrido: {elapsed_time:.2f} segundos")
        print("Conversión completada exitosamente")
    else:
        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"Fin: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end_time))}")
        print(f"Tiempo transcurrido: {elapsed_time:.2f} segundos")
        print("Error en la conversión")

if __name__ == "__main__":
    main()